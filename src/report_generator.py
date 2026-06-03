"""Generate reports for YOLO offline road-scene risk analysis."""

from __future__ import annotations

import html
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd


EVENT_COLUMNS = [
    "video_name",
    "frame_number",
    "timestamp_sec",
    "event_type",
    "risk_level",
    "detected_objects",
    "object_count",
    "brightness",
    "evidence_image",
    "tester_note",
]


def build_summary(
    video_name: str,
    total_frames_analyzed: int,
    events: list[dict[str, Any]],
    detected_object_counts: dict[str, int] | None = None,
    analysis_interval_sec: float = 1.0,
    model: str = "yolov8n.pt",
    confidence_threshold: float = 0.35,
) -> dict[str, Any]:
    """Build aggregate metadata for one video-analysis run."""
    risk_level_counts = Counter(event["risk_level"] for event in events)
    event_type_counts = Counter(event["event_type"] for event in events)
    return {
        "video_name": video_name,
        "total_frames_analyzed": total_frames_analyzed,
        "total_risk_events": len(events),
        "risk_level_counts": {
            "High": risk_level_counts.get("High", 0),
            "Medium": risk_level_counts.get("Medium", 0),
            "Low": risk_level_counts.get("Low", 0),
        },
        "event_type_counts": dict(event_type_counts),
        "detected_object_counts": detected_object_counts or {},
        "analysis_interval_sec": analysis_interval_sec,
        "model": model,
        "confidence_threshold": confidence_threshold,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }


def write_reports(events: list[dict[str, Any]], summary: dict[str, Any], output_dir: Path) -> dict[str, Path]:
    """Write CSV, Excel, JSON, and HTML reports."""
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "event_log.csv"
    excel_path = output_dir / "risk_events.xlsx"
    summary_path = output_dir / "summary.json"
    html_path = output_dir / "test_report.html"

    frame = pd.DataFrame(events, columns=EVENT_COLUMNS)
    frame.to_csv(csv_path, index=False)
    frame.to_excel(excel_path, index=False)
    _adjust_excel_columns(excel_path)

    summary["output_files"] = {
        "event_log_csv": str(csv_path).replace("\\", "/"),
        "risk_events_xlsx": str(excel_path).replace("\\", "/"),
        "summary_json": str(summary_path).replace("\\", "/"),
        "test_report_html": str(html_path).replace("\\", "/"),
    }
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    html_path.write_text(_render_html(frame, summary), encoding="utf-8")
    return {"csv": csv_path, "excel": excel_path, "summary": summary_path, "html": html_path}


def _adjust_excel_columns(excel_path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return

    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 48)
    workbook.save(excel_path)


def _render_count_items(counts: dict[str, int]) -> str:
    if not counts:
        return "<p>No events detected.</p>"
    return "".join(
        f"<div class=\"metric\"><span>{html.escape(str(name))}</span><strong>{count}</strong></div>"
        for name, count in sorted(counts.items())
    )


def _relative_evidence_path(evidence_image: str) -> str:
    if not evidence_image:
        return ""
    path = Path(evidence_image)
    if len(path.parts) >= 2 and path.parts[0] == "outputs":
        return str(Path(*path.parts[1:])).replace("\\", "/")
    return evidence_image.replace("\\", "/")


def _render_rows(frame: pd.DataFrame) -> str:
    if frame.empty:
        return "<tr><td colspan=\"8\">No risk events were detected.</td></tr>"

    rows = []
    for record in frame.to_dict(orient="records"):
        evidence = _relative_evidence_path(str(record.get("evidence_image") or ""))
        evidence_html = (
            f'<a href="{html.escape(evidence)}"><img src="{html.escape(evidence)}" alt="Evidence frame"></a>'
            if evidence
            else ""
        )
        rows.append(
            "<tr>"
            f"<td>{html.escape(str(record['frame_number']))}</td>"
            f"<td>{html.escape(str(record['timestamp_sec']))}</td>"
            f"<td>{html.escape(str(record['risk_level']))}</td>"
            f"<td>{html.escape(str(record['event_type']))}</td>"
            f"<td>{html.escape(str(record['detected_objects']))}</td>"
            f"<td>{html.escape(str(record['brightness']))}</td>"
            f"<td>{evidence_html}</td>"
            f"<td>{html.escape(str(record['tester_note']))}</td>"
            "</tr>"
        )
    return "\n".join(rows)


def _render_html(frame: pd.DataFrame, summary: dict[str, Any]) -> str:
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>YOLO Road Scene Risk Report</title>
  <style>
    body {{ margin: 0; color: #172026; background: #ffffff; font-family: "Segoe UI", Tahoma, sans-serif; }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 32px 20px 56px; }}
    header {{ border-bottom: 1px solid #d9e0e6; margin-bottom: 24px; padding-bottom: 18px; }}
    .summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin: 22px 0; }}
    .metric {{ background: #f6f8fa; border: 1px solid #d9e0e6; border-radius: 6px; padding: 14px; }}
    .metric span {{ color: #66727d; display: block; font-size: 0.9rem; }}
    .metric strong {{ display: block; font-size: 1.55rem; margin-top: 4px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 12px; font-size: 0.92rem; }}
    th, td {{ border-bottom: 1px solid #d9e0e6; padding: 10px; text-align: left; vertical-align: top; }}
    th {{ background: #f6f8fa; }}
    img {{ max-width: 180px; border: 1px solid #d9e0e6; border-radius: 4px; }}
  </style>
</head>
<body>
  <main>
    <header>
      <h1>YOLO Offline Road Scene Risk Report</h1>
      <p>Video analyzed: {html.escape(str(summary["video_name"]))}</p>
      <p>Generated at: {html.escape(str(summary["generated_at"]))}</p>
    </header>
    <section class="summary">
      <div class="metric"><span>Total frames analyzed</span><strong>{summary["total_frames_analyzed"]}</strong></div>
      <div class="metric"><span>Total risk events</span><strong>{summary["total_risk_events"]}</strong></div>
      <div class="metric"><span>Interval seconds</span><strong>{summary["analysis_interval_sec"]}</strong></div>
      <div class="metric"><span>Confidence threshold</span><strong>{summary["confidence_threshold"]}</strong></div>
      {_render_count_items(summary["risk_level_counts"])}
    </section>
    <section><h2>Event Type Counts</h2><div class="summary">{_render_count_items(summary["event_type_counts"])}</div></section>
    <section>
      <h2>Risk Event Log</h2>
      <table>
        <thead><tr><th>Frame</th><th>Time</th><th>Risk</th><th>Event</th><th>Objects</th><th>Brightness</th><th>Evidence</th><th>Note</th></tr></thead>
        <tbody>{_render_rows(frame)}</tbody>
      </table>
    </section>
  </main>
</body>
</html>
"""
